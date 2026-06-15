from __future__ import annotations

from datetime import datetime, time, timedelta
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from app.schemas import ParsedIssue, ParsedRoute, ParsedStop


HEADER_KEYWORDS = {
    "stop": ("остановочный пункт", "остановка", "оп"),
    "streets": ("наименования улиц", "улиц", "улица"),
    "latitude": ("широта",),
    "longitude": ("долгота",),
    "distance": ("расст. программн", "расст", "расстояние"),
    "cumulative_distance": ("общ программн", "общ", "протяженность"),
    "travel_time_between": ("время движения между оп", "между оп"),
    "cumulative_time": ("время движения нарастающ", "время движения", "нарастающ"),
}


def normalize_text(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").strip().lower().split())


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "").replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def time_to_minutes(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, timedelta):
        return round(value.total_seconds() / 60, 2)
    if isinstance(value, time):
        return float(value.hour * 60 + value.minute + round(value.second / 60, 2))
    if isinstance(value, datetime):
        return float(value.hour * 60 + value.minute + round(value.second / 60, 2))
    if isinstance(value, (int, float)):
        number = float(value)
        if 0 < number < 1:
            return round(number * 24 * 60, 2)
        return round(number, 2)
    text = str(value).strip()
    if ":" in text:
        parts = [int(float(part)) for part in text.split(":") if part != ""]
        if len(parts) == 2:
            return float(parts[0] * 60 + parts[1])
        if len(parts) == 3:
            return round(parts[0] * 60 + parts[1] + parts[2] / 60, 2)
    return to_float(text)


def parse_route_title(values: list[list[Any]], file_name: str) -> tuple[str | None, str | None]:
    for row in values:
        text = " ".join(str(cell) for cell in row if cell is not None)
        if "маршрут" in text.lower() and "№" in text:
            number_match = re.search(r"маршрут\s*№\s*([^\s\"“”]+)", text, flags=re.IGNORECASE)
            name_match = re.search(r"[\"“](.*?)[\"”]", text)
            number = normalize_route_number(number_match.group(1)) if number_match else None
            name = name_match.group(1).strip() if name_match else None
            return number, name
    return route_number_from_filename(file_name), None


def route_number_from_filename(file_name: str) -> str | None:
    stem = Path(file_name).stem
    match = re.search(r"[мm](\d{1,4})", stem, flags=re.IGNORECASE)
    if not match:
        match = re.search(r"(\d{1,4})", stem)
    return normalize_route_number(match.group(1)) if match else None


def normalize_route_number(value: Any) -> str | None:
    text = str(value).strip()
    digits = re.sub(r"\D", "", text)
    return str(int(digits)) if digits else text or None


def find_direction_rows(values: list[list[Any]]) -> tuple[int | None, int | None]:
    forward = backward = None
    for index, row in enumerate(values):
        row_text = normalize_text(" ".join(str(cell) for cell in row if cell is not None))
        if forward is None and "прямое направление" in row_text:
            forward = index
        if backward is None and "обратное направление" in row_text:
            backward = index
    return forward, backward


def find_header(values: list[list[Any]], start: int, end: int | None = None) -> tuple[int | None, dict[str, int]]:
    end = len(values) if end is None else min(end, len(values))
    best_row = None
    best_map: dict[str, int] = {}
    best_score = 0
    for row_index in range(max(0, start), end):
        row = values[row_index]
        mapping: dict[str, int] = {}
        mapping_scores: dict[str, int] = {}
        for col_index, cell in enumerate(row):
            text = normalize_text(cell)
            for key, keywords in HEADER_KEYWORDS.items():
                score = max((len(keyword) for keyword in keywords if keyword in text), default=0)
                if score and score > mapping_scores.get(key, 0):
                    mapping[key] = col_index
                    mapping_scores[key] = score
        score = len(mapping)
        if score > best_score:
            best_score = score
            best_row = row_index
            best_map = mapping
    if "stop" not in best_map:
        return None, best_map
    return best_row, best_map


def cell(row: list[Any], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def parse_direction(values: list[list[Any]], marker_row: int, end_row: int | None, direction: str) -> tuple[list[ParsedStop], list[ParsedIssue], float | None, float | None]:
    issues: list[ParsedIssue] = []
    header_row, columns = find_header(values, marker_row + 1, (end_row or len(values)))
    if header_row is None:
        header_row, columns = find_header(values, 0, len(values))
    if "cumulative_distance" not in columns:
        issues.append(ParsedIssue("error", "missing_length_column", "Не найдена колонка общ программн/протяженность."))
    if "cumulative_time" not in columns:
        issues.append(ParsedIssue("error", "missing_time_column", "Не найдена колонка время движения нарастающ."))
    if header_row is None:
        return [], issues, None, None

    stops: list[ParsedStop] = []
    length_values: list[float] = []
    time_values: list[float] = []
    end_row = len(values) if end_row is None else end_row
    for row_index in range(header_row + 1, end_row):
        row = values[row_index]
        row_text = normalize_text(" ".join(str(value) for value in row if value is not None))
        if "направление" in row_text or not any(value not in (None, "") for value in row):
            continue
        stop_name = str(cell(row, columns.get("stop")) or "").strip()
        if not stop_name:
            continue
        cumulative_distance = to_float(cell(row, columns.get("cumulative_distance")))
        cumulative_time = time_to_minutes(cell(row, columns.get("cumulative_time")))
        if cumulative_distance is not None:
            length_values.append(cumulative_distance)
        if cumulative_time is not None:
            time_values.append(cumulative_time)
        stops.append(
            ParsedStop(
                direction=direction,
                order_no=len(stops) + 1,
                stop_name=stop_name,
                streets=str(cell(row, columns.get("streets")) or "").strip() or None,
                latitude=to_float(cell(row, columns.get("latitude"))),
                longitude=to_float(cell(row, columns.get("longitude"))),
                distance_m=to_float(cell(row, columns.get("distance"))),
                cumulative_distance_m=cumulative_distance,
                travel_time_between_min=time_to_minutes(cell(row, columns.get("travel_time_between"))),
                cumulative_time_min=cumulative_time,
            )
        )
    length_km = round(max(length_values) / 1000, 2) if length_values else None
    trip_time = round(max(time_values), 2) if time_values else None
    return stops, issues, length_km, trip_time


def avg_speed(length_km: float | None, minutes: float | None) -> float | None:
    if not length_km or not minutes:
        return None
    return round(length_km / (minutes / 60), 2)


def validate_route(route: ParsedRoute) -> None:
    if not route.route_number:
        route.issues.append(ParsedIssue("error", "missing_route_number", "Не найден номер маршрута."))
    for direction, length, minutes, speed in (
        ("прямого", route.length_forward_km, route.trip_time_forward_min, route.avg_speed_forward_kmh),
        ("обратного", route.length_backward_km, route.trip_time_backward_min, route.avg_speed_backward_kmh),
    ):
        if not length:
            route.issues.append(ParsedIssue("error", "zero_length", f"Протяженность {direction} направления равна 0 или не найдена."))
        if not minutes:
            route.issues.append(ParsedIssue("error", "zero_trip_time", f"Время рейса {direction} направления равно 0 или не найдено."))
        if speed is not None and speed < 5:
            route.issues.append(ParsedIssue("warning", "speed_too_low", f"Средняя скорость {direction} направления менее 5 км/ч."))
        if speed is not None and speed > 80:
            route.issues.append(ParsedIssue("warning", "speed_too_high", f"Средняя скорость {direction} направления более 80 км/ч."))
    if route.length_forward_km and route.length_backward_km:
        diff = abs(route.length_forward_km - route.length_backward_km)
        base = max(route.length_forward_km, route.length_backward_km)
        if base and diff / base > 0.2:
            route.issues.append(ParsedIssue("warning", "direction_length_mismatch", "Большое расхождение протяженности прямого и обратного направления."))
    if any(issue.severity == "error" for issue in route.issues):
        route.data_status = "error"
    elif route.issues:
        route.data_status = "warning"
    else:
        route.data_status = "ok"
    route.comment = "; ".join(issue.message for issue in route.issues) or None


def parse_workbook(path: Path, original_path: str | None = None, modified_at: datetime | None = None) -> ParsedRoute:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet_name = next((name for name in workbook.sheetnames if name.lower() == "параметры"), None)
    file_path = original_path or str(path)
    if not sheet_name:
        route = ParsedRoute(None, None, path.name, file_path, None, modified_at)
        route.issues.append(ParsedIssue("error", "missing_parameters_sheet", "Отсутствует вкладка параметры."))
        validate_route(route)
        return route

    worksheet = workbook[sheet_name]
    values = [list(row) for row in worksheet.iter_rows(values_only=True)]
    route_number, route_name = parse_route_title(values, path.name)
    forward_row, backward_row = find_direction_rows(values)
    route = ParsedRoute(route_number, route_name, path.name, file_path, sheet_name, modified_at)
    if forward_row is None:
        route.issues.append(ParsedIssue("error", "missing_forward_direction", "Не найден блок прямое направление."))
    if backward_row is None:
        route.issues.append(ParsedIssue("error", "missing_backward_direction", "Не найден блок обратное направление."))

    if forward_row is not None:
        stops, issues, route.length_forward_km, route.trip_time_forward_min = parse_direction(values, forward_row, backward_row, "forward")
        route.stops.extend(stops)
        route.issues.extend(issues)
    if backward_row is not None:
        stops, issues, route.length_backward_km, route.trip_time_backward_min = parse_direction(values, backward_row, None, "backward")
        route.stops.extend(stops)
        route.issues.extend(issues)

    forward_stops = [stop for stop in route.stops if stop.direction == "forward"]
    backward_stops = [stop for stop in route.stops if stop.direction == "backward"]
    route.stops_forward_count = len(forward_stops)
    route.stops_backward_count = len(backward_stops)
    route.first_stop_forward = forward_stops[0].stop_name if forward_stops else None
    route.last_stop_forward = forward_stops[-1].stop_name if forward_stops else None
    route.first_stop_backward = backward_stops[0].stop_name if backward_stops else None
    route.last_stop_backward = backward_stops[-1].stop_name if backward_stops else None
    route.avg_speed_forward_kmh = avg_speed(route.length_forward_km, route.trip_time_forward_min)
    route.avg_speed_backward_kmh = avg_speed(route.length_backward_km, route.trip_time_backward_min)
    validate_route(route)
    return route
